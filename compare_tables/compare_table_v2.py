#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
表格比较脚本
用法: python compare_tables.py <表A文件> <表B文件>
"""

import pandas as pd
import re
import argparse
from pathlib import Path
from typing import List, Tuple
from openpyxl.styles import PatternFill


def decode_case_2(encoded_part: str) -> str:
    """
    实现情况2的复杂解码逻辑：将每两位十进制数解释为 ASCII 码，并转换回字符。
    """
    main_number = ""
    for k in range(0, len(encoded_part), 2):
        pair = encoded_part[k:k + 2]
        if pair.isdigit():
            try:
                ascii_code = int(pair)
                if 32 <= ascii_code <= 126:
                    main_number += chr(ascii_code)
            except ValueError:
                pass
    return main_number


def convert_base36(dec_value: int) -> str:
    """
    实现情况3的36进制转换逻辑：转换为大写36进制，并用 '0' 填充至 7 位。
    """
    if dec_value is None or dec_value < 0:
        return "0000000"
    
    base36 = ""
    BASE = 36
    BASE36_CHARS = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    
    if dec_value == 0:
        base36 = "0"
    else:
        temp_val = dec_value
        while temp_val > 0:
            base36 = BASE36_CHARS[temp_val % BASE] + base36
            temp_val //= BASE
            
    return base36.upper().zfill(7)


def process_encoded_data(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[Tuple[int, int, str]]]:
    """
    根据原始 Excel 脚本的逻辑处理 Pandas DataFrame 中的字符串数据。
    """
    df_processed = df.copy()
    df_processed.reset_index(drop=True, inplace=True) # 必需重置索引，否则 enumerate 的 i 与 loc[i] 不对应
    text_format_cells: List[Tuple[int, int, str]] = []
    changed_cols = set()
    
    target_col_idx = -1
    if '条码' in df_processed.columns:
         target_col_idx = df_processed.columns.get_loc('条码')

    for i, row in enumerate(df_processed.itertuples(index=False)):
        for j, cell_value in enumerate(row):
            # 如果找到了'条码'列，仅处理该列；否则处理所有列
            if target_col_idx != -1 and j != target_col_idx:
                continue

            if isinstance(cell_value, str):
                text = cell_value.strip()
                is_processed = False
                new_value = None
                                    
                # 情况3：20位且以 "C" 开头，且C后面跟2位数字
                if len(text) == 20 and text.startswith("C") and text[1:3].isdigit():
                    dec_part = text[5:16]
                    dec_value = int(dec_part) if dec_part.isdigit() else None
                    base36 = convert_base36(dec_value)
                    last_four = text[16:20]
                    new_value = f"FBA15{base36}U00{last_four}"
                    is_processed = True
                    changed_cols.add(j)
                
                # 更新处理后的值（不覆盖原值，新增列存放FBA条码）
                if is_processed and new_value is not None:
                    df_processed.loc[i, "fba条码"] = new_value
                else:
                    df_processed.loc[i, "fba条码"] = text
                
                if not is_processed and re.fullmatch(r'\d+', text):
                    text_format_cells.append((i, j, text))

    return df_processed, text_format_cells, changed_cols


def preprocess_scan_list(df: pd.DataFrame) -> pd.DataFrame:
    """按要求预处理：去掉末尾4行，补充缺少两个逗号的行，并按第2/4列去重。"""
    if df.empty:
        return df.copy()

    df_processed = df.iloc[:-4].copy() if len(df) > 4 else df.copy()
    col2 = df_processed.columns[1]
    col4 = df_processed.columns[3]

    extra_rows = []
    for idx, row in df_processed.iterrows():
        val = "" if pd.isna(row[col2]) else str(row[col2])
        comma_count = val.count(',')

        if comma_count == 2:
            cabinet_no, channel_no, pallet_no = val.split(',', 2)
            df_processed.loc[idx, "箱号"] = cabinet_no
            df_processed.loc[idx, "渠道号"] = channel_no
            df_processed.loc[idx, "托盘号"] = pallet_no
        # 如果不符合格式，则将第2列的值复制到第4列，因为可能是把条码扫成了箱码
        if comma_count < 2:
            df_processed.loc[idx, "箱号"] = val
            new_row = row.copy()
            new_row[col4] = row[col2]
            # 标记该行为特殊扫描
            new_row["箱号"] = '条码作为托盘贴扫描'
            extra_rows.append(new_row)

    if extra_rows:
        df_processed = pd.concat([df_processed, pd.DataFrame(extra_rows, columns=df_processed.columns)], ignore_index=True)

    df_processed = df_processed.drop_duplicates(subset=[col2, col4], keep='first')
    return df_processed


def preprocess_pkg_list(filename: str) -> pd.DataFrame:
    """预处理"包裹清单"sheet，将分组列展开为统一五列"""
    raw_pkg = pd.read_excel(filename, sheet_name="包裹清单", header=None)
    cont_name = Path(filename).stem
    cont_name = cont_name.split("操作分表")[0]
    row2 = raw_pkg.iloc[1]
    col_pre = [i for i, v in row2.items() if pd.notna(v) and "预报单号" in str(v)]
    col_tuo = [i for i, v in row2.items() if pd.notna(v) and "托盘序号" in str(v)]
    col_ref = [i for i, v in row2.items() if pd.notna(v) and "出库" in str(v)]
    col_bad = [i for i, v in row2.items() if pd.notna(v) and "破损/不可识别" in str(v)]
    row1 = raw_pkg.iloc[0]
    
    # 读取包裹列表用于卡派渠道的跟踪号替换
    # 改为一对多映射：一个预报单号对应多个跟踪号的列表
    tracking_map = {}
    try:
        parcel_list = pd.read_excel(filename, sheet_name="包裹列表", header=0)
        # 建立预报单号到跟踪号列表的映射
        platform_col = "Platform Order Ref.1\n平台单号1"
        track_col = "Track Nr.\n跟踪号"
        if platform_col in parcel_list.columns and track_col in parcel_list.columns:
            for _, row in parcel_list.iterrows():
                platform_ref = row[platform_col]
                track_nr = row[track_col]
                if pd.notna(platform_ref) and pd.notna(track_nr):
                    key = str(platform_ref).strip()
                    value = str(track_nr).strip()
                    if key not in tracking_map:
                        tracking_map[key] = []
                    tracking_map[key].append(value)
            
            total_mappings = sum(len(v) for v in tracking_map.values())
            print(f"已建立 {len(tracking_map)} 个预报单号到 {total_mappings} 个跟踪号的映射")
    except Exception as e:
        print(f"警告: 无法读取包裹列表，卡派渠道预报单号不会被替换: {e}")

    combined = []
    for idx_pre, idx_tuo, idx_ref, idx_bad in zip(col_pre, col_tuo, col_ref, col_bad):
        channel = row1.iloc[idx_pre]
        group_df = raw_pkg.iloc[2:, [idx_pre, idx_tuo, idx_ref, idx_bad]].copy()
        group_df.columns = ["预报单号", "托盘序号", "出库Ref", "破损/不可识别"]
        group_df = group_df.dropna(how="all")
        
        # 记录Excel位置信息（行号和各列的实际位置）
        # 行号从3开始（第1行渠道，第2行标题，第3行开始数据）
        # 列号：pandas索引从0开始，Excel列号从1开始，所以需要+1
        excel_row_numbers = list(range(3, 3 + len(group_df)))
        group_df['_excel_row'] = excel_row_numbers
        group_df['_excel_col_start'] = idx_pre + 1  # 预报单号列（Excel列号）
        
        # 查找"实际扫描"和"破损/不可识别"列的实际位置
        # 从当前渠道的预报单号列开始，向右查找
        scan_col = None
        damaged_col = None
        for offset in range(10):  # 假设一个渠道不超过10列
            check_idx = idx_pre + offset
            if check_idx < len(row2):
                header = row2.iloc[check_idx]
                if pd.notna(header):
                    if "实际扫描" in str(header) and scan_col is None:
                        scan_col = check_idx + 1  # Excel列号
                    elif ("破损" in str(header) or "不可识别" in str(header)) and damaged_col is None:
                        damaged_col = check_idx + 1  # Excel列号
        
        # 记录实际列位置
        group_df['_excel_col_scan'] = scan_col if scan_col else idx_pre + 4  # 默认+3
        group_df['_excel_col_damaged'] = damaged_col if damaged_col else idx_pre + 5  # 默认+4
        
        # 如果是卡派渠道，替换预报单号为跟踪号
        is_kapai = "卡派" in str(channel)
        if is_kapai and tracking_map:
            replaced_count = 0
            # 使用计数器记录每个预报单号已使用的索引
            forecast_usage_counter = {}
            
            def replace_with_tracking(x):
                nonlocal replaced_count
                if pd.notna(x):
                    key = str(x).strip()
                    if key in tracking_map:
                        # 获取该预报单号当前使用的索引
                        if key not in forecast_usage_counter:
                            forecast_usage_counter[key] = 0
                        
                        idx = forecast_usage_counter[key]
                        tracking_list = tracking_map[key]
                        
                        # 如果索引超出列表范围，循环使用（或使用最后一个）
                        if idx < len(tracking_list):
                            result = tracking_list[idx]
                        else:
                            # 如果跟踪号不够用，使用最后一个
                            result = tracking_list[-1]
                        
                        # 增加计数器
                        forecast_usage_counter[key] += 1
                        replaced_count += 1
                        return result
                return x
            
            group_df["预报单号"] = group_df["预报单号"].apply(replace_with_tracking)
            print(f"卡派渠道 '{channel}': 已替换 {replaced_count}/{len(group_df)} 个预报单号为跟踪号")
        
        group_df.insert(4, '箱号', cont_name)
        group_df.insert(5, "渠道号", channel)
        combined.append(group_df)

    combined_df = pd.concat(combined, ignore_index=True) if combined else pd.DataFrame(
        columns=["渠道号", "预报单号", "托盘序号", "出库ref", "破损/不可识别"])
    combined_df = combined_df.dropna(how="all")
    return combined_df


def compare_tables(df_a: pd.DataFrame, df_b: pd.DataFrame) -> pd.DataFrame:
    """
    表A：包含条码、箱号、渠道号（条码与B的预报单号/托盘序号匹配 --- 扫描表
    表B：包含预报单号、托盘序号、箱号、渠道号 --- 操作分表
    返回：在表B的副本上添加对齐结果与扫描信息
    """
    res = df_b.copy()
    
    # 我们将构建一个新的行列表，因为一行可能对应多个匹配
    new_rows = []
    
    for _, row in res.iterrows():
        tuo = row.get("托盘序号")
        pre = row.get("预报单号")
        key = f"{'' if pd.isna(tuo) else str(tuo).strip()}{'' if pd.isna(pre) else str(pre).strip()}"

        # 默认字段初始化
        row_base = row.copy()
        row_base['条码匹配'] = ''
        row_base["箱号对齐"] = ""
        row_base["渠道对齐"] = ""
        row_base["扫描箱号"] = ""
        row_base["扫描渠道号"] = ""
        row_base["扫描托盘号"] = ""
        row_base['原始扫描序号'] = ''

        if key == "":
            new_rows.append(row_base)
            continue

        mask = df_a.get("fba条码").astype(str).apply(lambda x: x != "" and len(x) > 10 and x in key)
        match_rows = df_a[mask]
        
        if match_rows.empty:
            row_base['条码匹配'] = '否'
            new_rows.append(row_base)
            continue
        
        # 遍历所有匹配行
        for _, match in match_rows.iterrows():
            current_row = row_base.copy()
            
            scan_box = match.get("箱号")
            scan_channel = match.get("渠道号")
            scan_pallet = match.get("托盘号")
            scan_ori = match.get("条码")
            
            b_box = row.get("箱号")
            b_channel = row.get("渠道号")

            same_box = pd.notna(scan_box) and pd.notna(b_box) and str(scan_box) == str(b_box)
            same_channel = pd.notna(scan_channel) and pd.notna(b_channel) and str(scan_channel) == str(b_channel)

            current_row["箱号对齐"] = "是" if same_box else "否"
            current_row["渠道对齐"] = "是" if same_channel else "否"
            current_row["原始扫描序号"] = scan_ori

            if not same_box:
                if pd.isna(scan_box) or str(scan_box).strip() == "":
                    current_row["扫描箱号"] = "扫描箱码格式不符"
                else:
                    current_row["扫描箱号"] = scan_box
            if not same_channel:
                current_row["扫描渠道号"] = scan_channel
            
            # 始终显示扫描的托盘号
            current_row["扫描托盘号"] = scan_pallet

            new_rows.append(current_row)

    return pd.DataFrame(new_rows)



def export_with_colors(df: pd.DataFrame, filename: str):
    """
    导出DataFrame到Excel，并根据条件对行进行颜色标记：
    - 条码匹配=否：整行标红
    - 箱号对齐或渠道对齐只有一个=否：标黄
    - 箱号和渠道都对齐：标绿
    """
    required_cols = ['预报单号', '托盘序号', '出库Ref', '破损/不可识别', '箱号', '渠道号', 
                     '条码匹配', '箱号对齐', '渠道对齐', '扫描箱号', '扫描渠道号', '扫描托盘号', '原始扫描序号']
    
    available_cols = [col for col in required_cols if col in df.columns]
    export_df = df[available_cols].copy()
    export_df = export_df.reset_index(drop=True)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='Sheet1', index=False)
        worksheet = writer.sheets['Sheet1']
        
        red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
        yellow_fill = PatternFill(start_color="EEFF00", end_color="EEFF00", fill_type="solid")
        green_fill = PatternFill(start_color="14E01E", end_color="14E01E", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFA500", fill_type="solid")
        
        # 查重逻辑：对‘原始扫描序号’列进行查重（忽略空值）
        try:
            # 确保转为字符串处理，去除空白
            original_scan_col = export_df['原始扫描序号'].astype(str).apply(lambda x: x.strip())
            # 排除空值、nan、None等
            valid_mask = (original_scan_col != '') & (original_scan_col != 'nan') & (original_scan_col != 'None')
            valid_series = original_scan_col[valid_mask]
            
            # 找出重复值的索引
            # keep=False 表示标记所有重复出现的值（包括第一次出现的）
            duplicates = valid_series[valid_series.duplicated(keep=False)]
            duplicate_indices = set(duplicates.index)
        except Exception as e:
            print(f"查重逻辑出错: {e}")
            duplicate_indices = set()

        for idx, row in export_df.iterrows():
            row_num = idx + 2
            
            code_match = str(row.get('条码匹配', '')).strip()
            box_align = str(row.get('箱号对齐', '')).strip()
            channel_align = str(row.get('渠道对齐', '')).strip()
            
            fill_color = None
            if code_match == '否':
                fill_color = red_fill
            elif box_align == '是' and channel_align == '是':
                fill_color = green_fill
            elif box_align == '否' or channel_align == '否':
                fill_color = yellow_fill
            
            # 如果是重复的原始扫描序号，覆盖为橙色
            if idx in duplicate_indices:
                fill_color = orange_fill
            
            if fill_color:
                for col_idx in range(1, len(available_cols) + 1):
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.fill = fill_color
    
    print(f"已导出到 {filename}")


def export_backfill_to_original(original_file: str, compared_df: pd.DataFrame, output_filename: str):
    """
    将比对结果回填到原始Excel文件，保留原始格式
    使用DataFrame中记录的Excel位置信息(_excel_row, _excel_col_start)直接回填
    - 在"破损/不可识别"列填入：扫描箱号,扫描渠道号（如果不对齐）
    - 在"实际扫描"列填入：原始扫描序号（如果匹配成功）
    - 填充背景色：红/黄/绿
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    # 加载原始工作簿
    wb = load_workbook(original_file)
    ws = wb["包裹清单"]
    
    # 定义颜色
    red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
    yellow_fill = PatternFill(start_color="EEFF00", end_color="EEFF00", fill_type="solid")
    green_fill = PatternFill(start_color="14E01E", end_color="14E01E", fill_type="solid")
    
    # 使用记录的实际列位置进行回填
    # _excel_col_scan 和 _excel_col_damaged 是实际的列位置
    
    # 统计
    processed_count = 0
    filled_color_count = 0
    
    # 遍历比对结果DataFrame并回填数据
    for idx, row in compared_df.iterrows():
        # 获取Excel位置信息
        excel_row = row.get('_excel_row')
        col_scan = row.get('_excel_col_scan')  # 实际扫描列的实际位置
        col_damaged = row.get('_excel_col_damaged')  # 破损/不可识别列的实际位置
        
        if pd.isna(excel_row) or pd.isna(col_scan) or pd.isna(col_damaged):
            continue
        
        excel_row = int(excel_row)
        col_scan = int(col_scan)
        col_damaged = int(col_damaged)
        
        # 填充"破损/不可识别"列
        scan_box = row.get('扫描箱号', '')
        scan_channel = row.get('扫描渠道号', '')
        if pd.notna(scan_box) or pd.notna(scan_channel):
            box_str = str(scan_box) if pd.notna(scan_box) else ''
            channel_str = str(scan_channel) if pd.notna(scan_channel) else ''
            # 智能拼接
            if box_str and channel_str:
                damaged_value = f"{box_str},{channel_str}"
            elif box_str:
                damaged_value = box_str
            elif channel_str:
                damaged_value = channel_str
            else:
                damaged_value = None
            
            if damaged_value:
                ws.cell(row=excel_row, column=col_damaged).value = damaged_value
        
        # 填充"实际扫描"列
        scan_ori = row.get('原始扫描序号', '')
        if pd.notna(scan_ori) and scan_ori:
            ws.cell(row=excel_row, column=col_scan).value = str(scan_ori)
        
        # 填充颜色
        code_match = str(row.get('条码匹配', '')).strip()
        box_align = str(row.get('箱号对齐', '')).strip()
        channel_align = str(row.get('渠道对齐', '')).strip()
        
        # 判断是否需要填充颜色（只要有扫描数据就填充）
        has_scan_info = pd.notna(row.get('扫描箱号')) or pd.notna(row.get('扫描渠道号')) or pd.notna(row.get('原始扫描序号'))
        
        if not has_scan_info:
            fill_color = None
        elif code_match == '否':
            fill_color = red_fill
        elif box_align == '是' and channel_align == '是':
            fill_color = green_fill
        else:
            # 默认黄色（有扫描数据但不是完全对齐）
            fill_color = yellow_fill
        
        if fill_color:
            # 给"实际扫描"和"破损/不可识别"列上色（无论是否有值）
            # 这样可以通过颜色看出匹配状态：
            # - 红色：条码未匹配（空单元格）
            # - 黄色：部分匹配（可能有数据或部分为空）
            # - 绿色：完全匹配（通常有数据）
            ws.cell(row=excel_row, column=col_scan).fill = fill_color
            ws.cell(row=excel_row, column=col_damaged).fill = fill_color
            
            filled_color_count += 1
        
        processed_count += 1
    
    # 保存到新文件
    wb.save(output_filename)
    print(f"  回填统计: 处理 {processed_count} 行, 填充颜色 {filled_color_count} 行")
    print(f"已回填到 {output_filename}")


def process_single_file(table_a_path, table_b_path: str, preprocessed_scan_df=None) -> bool:
    """
    处理单个文件对的比对
    table_a_path: 扫描数据路径（str）或已处理的DataFrame
    table_b_path: 包裹清单路径
    preprocessed_scan_df: 可选，已预处理的扫描数据DataFrame
    返回: True表示成功处理，False表示跳过
    """
    # 生成输出文件名并检查是否已存在
    table_b_name = Path(table_b_path).stem
    output_dir = Path("compare_tables_test/output")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_filename = output_dir / f"比较结果_{table_b_name}.xlsx"
    backfill_filename = output_dir / f"回填结果_{table_b_name}.xlsx"
    
    if output_filename.exists() or backfill_filename.exists():
        print(f"跳过: 输出文件已存在")
        if output_filename.exists():
            print(f"  - {output_filename}")
        if backfill_filename.exists():
            print(f"  - {backfill_filename}")
        return False
    
    # 读取表A（扫描数据）- 如果没有提供预处理的数据
    if preprocessed_scan_df is None:
        print(f"正在读取表A: {table_a_path}")
        scan_list = pd.read_excel(table_a_path)
        # 1. 先预处理
        preprocessed_list = preprocess_scan_list(scan_list)
        # 2. 再解码
        preprocessed_scan_df, _, _ = process_encoded_data(preprocessed_list)
        print(f"表A预处理完成，共 {len(preprocessed_scan_df)} 行")
    else:
        print(f"使用已合并的扫描数据，共 {len(preprocessed_scan_df)} 行")
    
    # 读取表B（包裹清单）
    print(f"正在读取表B: {table_b_path}")
    raw_pkg2 = preprocess_pkg_list(table_b_path)
    print(f"表B预处理完成，共 {len(raw_pkg2)} 行")
    
    # 比较表格
    print("正在比较表格...")
    compared_df = compare_tables(preprocessed_scan_df, raw_pkg2)
    
    # 导出结果1：比较结果（包含所有比对列）
    print(f"正在导出比较结果到: {output_filename}")
    export_with_colors(compared_df, str(output_filename))
    
    # 导出结果2：回填到原始文件（保留原格式）
    backfill_filename = output_dir / f"回填结果_{table_b_name}.xlsx"
    print(f"正在导出回填结果到: {backfill_filename}")
    export_backfill_to_original(table_b_path, compared_df, str(backfill_filename))
    
    print("完成！")
    return True


def load_scan_data(table_a_path: Path) -> pd.DataFrame:
    """
    加载并预处理扫描数据（支持文件夹自动合并或单文件读取）
    """
    preprocessed_scan_df = None
    
    if table_a_path.is_dir():
        # 扫描数据文件夹模式：合并所有扫描文件
        print(f"检测到扫描数据文件夹: {table_a_path}")
        # 排除临时文件和Excel临时文件
        scan_files = sorted([f for f in table_a_path.glob("*.xlsx") 
                            if not f.name.startswith("~$") and not f.name.startswith("_merged")])
        
        if not scan_files:
            print("错误: 扫描数据文件夹中没有找到 .xlsx 文件")
            return None
        
        print(f"找到 {len(scan_files)} 个扫描文件:")
        for f in scan_files:
            print(f"  - {f.name}")
        
        # 合并所有扫描文件
        all_scan_data = []
        failed_files = []
        
        for scan_file in scan_files:
            print(f"正在读取: {scan_file.name}")
            try:
                scan_list = pd.read_excel(scan_file)
                # 调整顺序：先不解码，仅读取
                all_scan_data.append(scan_list)
                print(f"  ✓ 成功读取 {len(scan_list)} 行")
            except Exception as e:
                failed_files.append((scan_file.name, str(e)[:200]))
                print(f"  ✗ 读取失败: {str(e)[:100]}")
        
        # 检查是否有失败的文件
        if failed_files:
            print(f"\n错误: 有 {len(failed_files)} 个文件读取失败:")
            for filename, error in failed_files:
                print(f"  - {filename}: {error}")
            print("\n请修复这些文件后重试，或从input_scan文件夹中移除它们")
            return None
        
        # 合并并去重
        if not all_scan_data:
            print("错误: 没有成功读取任何扫描文件")
            return None
        
        print(f"\n成功读取所有 {len(scan_files)} 个扫描文件")
        combined_scan = pd.concat(all_scan_data, ignore_index=True)
        print(f"合并后: {len(combined_scan)} 行")
        
        # 1. 先进行预处理（清洗、去重）
        preprocessed_combined = preprocess_scan_list(combined_scan)
        print(f"清洗去重后: {len(preprocessed_combined)} 行")
        
        # 2. 再进行解码
        preprocessed_scan_df, _, _ = process_encoded_data(preprocessed_combined)
        print(f"扫描数据解码完成\n")
        
    elif table_a_path.exists():
        # 单文件模式
        if str(table_a_path) == 'compare_tables_test/input_scan':
             pass # default dummy, usually won't hit if exists check passed unless it is a file named that
        else:
             print(f"使用扫描数据文件: {table_a_path}")
             
        print(f"正在读取表A: {table_a_path}")
        try:
            scan_list = pd.read_excel(table_a_path)
            # 1. 先预处理
            preprocessed_list = preprocess_scan_list(scan_list)
            # 2. 再解码
            preprocessed_scan_df, _, _ = process_encoded_data(preprocessed_list)
            print(f"表A预处理完成，共 {len(preprocessed_scan_df)} 行")
        except Exception as e:
            print(f"读取失败: {e}")
            return None
    else:
        # Fallback for default string if not exists (though usually it should exist or be dir)
        print(f"错误: 找不到扫描数据 {table_a_path}")
        return None

    return preprocessed_scan_df


def main():
    parser = argparse.ArgumentParser(description='比较两个表格并导出结果')
    parser.add_argument('table_a', nargs='?', 
                        default='compare_tables_test/input_scan',
                        help='表A文件路径或文件夹（扫描数据），默认: compare_tables_test/input_scan (自动合并文件夹中所有扫描文件)')
    parser.add_argument('table_b', nargs='?',
                        default='compare_tables_test/input_pkg',
                        help='表B文件路径或文件夹（包裹清单），默认: compare_tables_test/input_pkg (批量处理文件夹)')
    
    args = parser.parse_args()
    
    # 1. 加载扫描数据
    table_a_path = Path(args.table_a)
    preprocessed_scan_df = load_scan_data(table_a_path)
    
    if preprocessed_scan_df is None:
        return

    # 2. 处理表B（批量或单文件）
    if args.table_b == 'compare_tables_test/input_pkg':
        print("使用默认包裹清单文件夹（批量处理模式）")
    
    table_b_path = Path(args.table_b)
    
    # 检查table_b是文件还是文件夹
    if table_b_path.is_dir():
        # 批量处理模式
        print(f"检测到文件夹，开始批量处理: {table_b_path}")
        # 过滤掉Excel临时文件
        xlsx_files = sorted([f for f in table_b_path.glob("*.xlsx") if not f.name.startswith("~$")])
        
        if not xlsx_files:
            print(f"错误: 文件夹中没有找到 .xlsx 文件")
            return
        
        print(f"找到 {len(xlsx_files)} 个文件")
        processed_count = 0
        skipped_count = 0
        
        for idx, xlsx_file in enumerate(xlsx_files, 1):
            print(f"[{idx}/{len(xlsx_files)}] 处理文件: {xlsx_file.name}")
            print("=" * 60)
            # 始终传递预处理好的df
            result = process_single_file(None, str(xlsx_file), preprocessed_scan_df=preprocessed_scan_df)
            
            if result:
                processed_count += 1
            else:
                skipped_count += 1
        
        print("\n" + "=" * 60)
        print(f"批量处理完成！")
        print(f"成功处理: {processed_count} 个文件")
        print(f"跳过: {skipped_count} 个文件")
        print(f"总计: {len(xlsx_files)} 个文件")
    else:
        # 单文件处理模式
        if not table_b_path.exists():
            print(f"错误: 文件不存在 - {table_b_path}")
            return
        
        # 始终传递预处理好的df
        process_single_file(None, args.table_b, preprocessed_scan_df=preprocessed_scan_df)


if __name__ == "__main__":
    main()

