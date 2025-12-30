#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import pandas as pd
import re
from pathlib import Path
from typing import List, Tuple
from openpyxl.styles import PatternFill

# =================================================================================================
#  Utility Functions (Originally from compare_utils.py)
# =================================================================================================

def decode_case_2(encoded_part: str) -> str:
    """
    实现情况2的复杂解码逻辑：将每两位十进制数解释为 ASCII 码，并转换回字符。
    """
    main_number = ""
    for k in range(0, len(encoded_part), 2):
        pair = encoded_part[k:k + 2]
        if pair.isdigit():
            try:
                ascii_code = int(pair)
                if 32 <= ascii_code <= 126:
                    main_number += chr(ascii_code)
            except ValueError:
                pass
    return main_number


def convert_base36(dec_value: int) -> str:
    """
    实现情况3的36进制转换逻辑：转换为大写36进制，并用 '0' 填充至 7 位。
    """
    if dec_value is None or dec_value < 0:
        return "0000000"
    
    base36 = ""
    BASE = 36
    BASE36_CHARS = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    
    if dec_value == 0:
        base36 = "0"
    else:
        temp_val = dec_value
        while temp_val > 0:
            base36 = BASE36_CHARS[temp_val % BASE] + base36
            temp_val //= BASE
            
    return base36.upper().zfill(7)


def process_encoded_data(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[Tuple[int, int, str]]]:
    """
    根据原始 Excel 脚本的逻辑处理 Pandas DataFrame 中的字符串数据。
    """
    df_processed = df.copy()
    df_processed.reset_index(drop=True, inplace=True)
    text_format_cells: List[Tuple[int, int, str]] = []
    changed_cols = set()
    
    target_col_idx = -1
    if '条码' in df_processed.columns:
         target_col_idx = df_processed.columns.get_loc('条码')

    for i, row in enumerate(df_processed.itertuples(index=False)):
        for j, cell_value in enumerate(row):
            # 如果找到了'条码'列，仅处理该列；否则处理所有列
            if target_col_idx != -1 and j != target_col_idx:
                continue

            if isinstance(cell_value, str):
                text = cell_value.strip()
                is_processed = False
                new_value = None
                                    
                # 情况3：20位且以 "C" 开头，且C后面跟2位数字
                if len(text) == 20 and text.startswith("C") and text[1:3].isdigit():
                    dec_part = text[5:16]
                    dec_value = int(dec_part) if dec_part.isdigit() else None
                    base36 = convert_base36(dec_value)
                    last_four = text[16:20]
                    new_value = f"FBA15{base36}U00{last_four}"
                    is_processed = True
                    changed_cols.add(j)
                
                # 更新处理后的值（不覆盖原值，新增列存放FBA条码）
                if is_processed and new_value is not None:
                    df_processed.loc[i, "fba条码"] = new_value
                else:
                    df_processed.loc[i, "fba条码"] = text
                
                if not is_processed and re.fullmatch(r'\d+', text):
                    text_format_cells.append((i, j, text))

    return df_processed, text_format_cells, changed_cols


def preprocess_scan_list(df: pd.DataFrame) -> pd.DataFrame:
    """按要求预处理：去掉末尾4行，补充缺少两个逗号的行，并按第2/4列去重。"""
    if df.empty:
        return df.copy()

    df_processed = df.iloc[:-4].copy() if len(df) > 4 else df.copy()
    col2 = df_processed.columns[1]
    col4 = df_processed.columns[3]

    extra_rows = []
    for idx, row in df_processed.iterrows():
        val = "" if pd.isna(row[col2]) else str(row[col2])
        comma_count = val.count(',')

        if comma_count == 2:
            cabinet_no, channel_no, pallet_no = val.split(',', 2)
            df_processed.loc[idx, "箱号"] = cabinet_no
            df_processed.loc[idx, "渠道号"] = channel_no
            df_processed.loc[idx, "托盘号"] = pallet_no
        # 如果不符合格式，则将第2列的值复制到第4列，因为可能是把条码扫成了箱码
        if comma_count < 2:
            df_processed.loc[idx, "箱号"] = val
            new_row = row.copy()
            new_row[col4] = row[col2]
            # 标记该行为特殊扫描
            new_row["箱号"] = '条码作为托盘贴扫描'
            extra_rows.append(new_row)

    if extra_rows:
        df_processed = pd.concat([df_processed, pd.DataFrame(extra_rows, columns=df_processed.columns)], ignore_index=True)

    df_processed = df_processed.drop_duplicates(subset=[col2, col4], keep='first')
    
    return df_processed


def preprocess_pkg_list(filename: str) -> pd.DataFrame:
    """预处理"包裹清单"sheet，将分组列展开为统一五列"""
    raw_pkg = pd.read_excel(filename, sheet_name="包裹清单", header=None)
    cont_name = Path(filename).stem
    cont_name = cont_name.split("操作分表")[0]
    row2 = raw_pkg.iloc[1]
    col_pre = [i for i, v in row2.items() if pd.notna(v) and "预报单号" in str(v)]
    col_tuo = [i for i, v in row2.items() if pd.notna(v) and "托盘序号" in str(v)]
    col_ref = [i for i, v in row2.items() if pd.notna(v) and "出库" in str(v)]
    col_bad = [i for i, v in row2.items() if pd.notna(v) and "破损/不可识别" in str(v)]
    row1 = raw_pkg.iloc[0]
    
    # 读取包裹列表用于卡派渠道的跟踪号替换
    tracking_map = {}
    try:
        parcel_list = pd.read_excel(filename, sheet_name="包裹列表", header=0)
        platform_col = "Platform Order Ref.1\n平台单号1"
        track_col = "Track Nr.\n跟踪号"
        if platform_col in parcel_list.columns and track_col in parcel_list.columns:
            for _, row in parcel_list.iterrows():
                platform_ref = row[platform_col]
                track_nr = row[track_col]
                if pd.notna(platform_ref) and pd.notna(track_nr):
                    key = str(platform_ref).strip()
                    value = str(track_nr).strip()
                    if key not in tracking_map:
                        tracking_map[key] = []
                    tracking_map[key].append(value)
            
            total_mappings = sum(len(v) for v in tracking_map.values())
            print(f"已建立 {len(tracking_map)} 个预报单号到 {total_mappings} 个跟踪号的映射")
    except Exception as e:
        print(f"警告: 无法读取包裹列表，卡派渠道预报单号不会被替换: {e}")

    combined = []
    for idx_pre, idx_tuo, idx_ref, idx_bad in zip(col_pre, col_tuo, col_ref, col_bad):
        channel = str(row1.iloc[idx_pre]).strip()
        # 去除前缀 logic moved here globally
        if channel.startswith("CWE-"):
            channel = channel[4:] # len("CWE-") = 4
        if channel.startswith("卡派-"):
            channel = channel[3:] # len("卡派-") = 3

        group_df = raw_pkg.iloc[2:, [idx_pre, idx_tuo, idx_ref, idx_bad]].copy()
        group_df.columns = ["预报单号", "托盘序号", "出库Ref", "破损/不可识别"]
        group_df = group_df.dropna(how="all")
        
        excel_row_numbers = list(range(3, 3 + len(group_df)))
        group_df['_excel_row'] = excel_row_numbers
        group_df['_excel_col_start'] = idx_pre + 1
        
        scan_col = None
        damaged_col = None
        for offset in range(10):
            check_idx = idx_pre + offset
            if check_idx < len(row2):
                header = row2.iloc[check_idx]
                if pd.notna(header):
                    if "实际扫描" in str(header) and scan_col is None:
                        scan_col = check_idx + 1
                    elif ("破损" in str(header) or "不可识别" in str(header)) and damaged_col is None:
                        damaged_col = check_idx + 1
        
        group_df['_excel_col_scan'] = scan_col if scan_col else idx_pre + 4
        group_df['_excel_col_damaged'] = damaged_col if damaged_col else idx_pre + 5
        
        # 使用原始 channel 字符串判断是否为卡派，因为 channel 变量可能已经被去除了前缀
        original_channel = str(row1.iloc[idx_pre])
        is_kapai = "卡派" in original_channel
        
        if is_kapai and tracking_map:
            replaced_count = 0
            forecast_usage_counter = {}
            
            def replace_with_tracking(x):
                nonlocal replaced_count
                if pd.notna(x):
                    key = str(x).strip()
                    if key in tracking_map:
                        if key not in forecast_usage_counter:
                            forecast_usage_counter[key] = 0
                        
                        idx = forecast_usage_counter[key]
                        tracking_list = tracking_map[key]
                        
                        if idx < len(tracking_list):
                            result = tracking_list[idx]
                        else:
                            result = tracking_list[-1]
                        
                        forecast_usage_counter[key] += 1
                        replaced_count += 1
                        return result
                return x
            
            group_df["预报单号"] = group_df["预报单号"].apply(replace_with_tracking)
            print(f"卡派渠道 '{channel}': 已替换 {replaced_count}/{len(group_df)} 个预报单号为跟踪号")
        
        group_df.insert(4, '箱号', cont_name)
        group_df.insert(5, "渠道号", channel)
        combined.append(group_df)

    combined_df = pd.concat(combined, ignore_index=True) if combined else pd.DataFrame(
        columns=["渠道号", "预报单号", "托盘序号", "出库ref", "破损/不可识别"])
    combined_df = combined_df.dropna(how="all")
    return combined_df


def compare_tables(df_a: pd.DataFrame, df_b: pd.DataFrame) -> pd.DataFrame:
    """
    比较表格逻辑
    """
    res = df_b.copy()
    new_rows = []
    
    for _, row in res.iterrows():
        tuo = row.get("托盘序号")
        pre = row.get("预报单号")
        key = f"{'' if pd.isna(tuo) else str(tuo).strip()}{'' if pd.isna(pre) else str(pre).strip()}"

        row_base = row.copy()
        row_base['条码匹配'] = ''
        row_base["箱号对齐"] = ""
        row_base["渠道对齐"] = ""
        row_base["扫描箱号"] = ""
        row_base["扫描渠道号"] = ""
        row_base["扫描托盘号"] = ""
        row_base['原始扫描序号'] = ''

        if key == "":
            new_rows.append(row_base)
            continue

        mask = df_a.get("fba条码").astype(str).apply(lambda x: x != "" and len(x) > 10 and x in key)
        match_rows = df_a[mask]
        
        if match_rows.empty:
            row_base['条码匹配'] = '否'
            new_rows.append(row_base)
            continue
        
        for _, match in match_rows.iterrows():
            current_row = row_base.copy()
            
            scan_box = match.get("箱号")
            scan_channel = match.get("渠道号")
            scan_pallet = match.get("托盘号")
            scan_ori = match.get("条码")
            
            b_box = row.get("箱号")
            b_channel = row.get("渠道号")

            same_box = pd.notna(scan_box) and pd.notna(b_box) and str(scan_box) == str(b_box)
            same_channel = pd.notna(scan_channel) and pd.notna(b_channel) and str(scan_channel) == str(b_channel)

            current_row["箱号对齐"] = "是" if same_box else "否"
            current_row["渠道对齐"] = "是" if same_channel else "否"
            current_row["原始扫描序号"] = scan_ori

            if not same_box:
                if pd.isna(scan_box) or str(scan_box).strip() == "":
                    current_row["扫描箱号"] = "扫描箱码格式不符"
                else:
                    current_row["扫描箱号"] = scan_box
            if not same_channel:
                current_row["扫描渠道号"] = scan_channel
            
            current_row["扫描托盘号"] = scan_pallet

            new_rows.append(current_row)

    return pd.DataFrame(new_rows)


def load_scan_data(table_a_path: Path) -> pd.DataFrame:
    """
    加载并预处理扫描数据
    """
    preprocessed_scan_df = None
    
    if table_a_path.is_dir():
        print(f"检测到扫描数据文件夹: {table_a_path}")
        scan_files = sorted([f for f in table_a_path.glob("*.xlsx") 
                            if not f.name.startswith("~$") and not f.name.startswith("_merged")])
        
        if not scan_files:
            print("错误: 扫描数据文件夹中没有找到 .xlsx 文件")
            return None
        
        print(f"找到 {len(scan_files)} 个扫描文件")
        
        all_scan_data = []
        for scan_file in scan_files:
            print(f"正在读取: {scan_file.name}")
            try:
                scan_list = pd.read_excel(scan_file)
                all_scan_data.append(scan_list)
            except Exception as e:
                print(f"  ✗ 读取失败: {str(e)[:100]}")
        
        if not all_scan_data:
            print("错误: 没有成功读取任何扫描文件")
            return None
        
        combined_scan = pd.concat(all_scan_data, ignore_index=True)
        # 1. preprocess (clean/dedupe)
        preprocessed_combined = preprocess_scan_list(combined_scan)
        # 2. decode
        preprocessed_scan_df, _, _ = process_encoded_data(preprocessed_combined)
        print(f"扫描数据预处理完成，共 {len(preprocessed_scan_df)} 行\n")
        
    elif table_a_path.exists():
        print(f"使用扫描数据文件: {table_a_path}")
        try:
            scan_list = pd.read_excel(table_a_path)
            # 1. preprocess
            preprocessed_list = preprocess_scan_list(scan_list)
            # 2. decode
            preprocessed_scan_df, _, _ = process_encoded_data(preprocessed_list)
            print(f"表A预处理完成，共 {len(preprocessed_scan_df)} 行")
        except Exception as e:
            print(f"读取失败: {e}")
            return None
    else:
        print(f"错误: 找不到扫描数据 {table_a_path}")
        return None

    return preprocessed_scan_df


def export_with_colors(df: pd.DataFrame, filename: str):
    """
    导出比较结果 excel
    """
    required_cols = ['预报单号', '托盘序号', '出库Ref', '破损/不可识别', '箱号', '渠道号', 
                     '条码匹配', '箱号对齐', '渠道对齐', '扫描箱号', '扫描渠道号', '扫描托盘号', '原始扫描序号']
    
    available_cols = [col for col in required_cols if col in df.columns]
    export_df = df[available_cols].copy()
    export_df = export_df.reset_index(drop=True)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='Sheet1', index=False)
        worksheet = writer.sheets['Sheet1']
        
        red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
        yellow_fill = PatternFill(start_color="EEFF00", end_color="EEFF00", fill_type="solid")
        green_fill = PatternFill(start_color="14E01E", end_color="14E01E", fill_type="solid")
        orange_fill = PatternFill(start_color="FFC000", end_color="FFA500", fill_type="solid")
        
        try:
            original_scan_col = export_df['原始扫描序号'].astype(str).apply(lambda x: x.strip())
            valid_mask = (original_scan_col != '') & (original_scan_col != 'nan') & (original_scan_col != 'None')
            valid_series = original_scan_col[valid_mask]
            duplicates = valid_series[valid_series.duplicated(keep=False)]
            duplicate_indices = set(duplicates.index)
        except Exception as e:
            print(f"查重逻辑出错: {e}")
            duplicate_indices = set()

        for idx, row in export_df.iterrows():
            row_num = idx + 2
            code_match = str(row.get('条码匹配', '')).strip()
            box_align = str(row.get('箱号对齐', '')).strip()
            channel_align = str(row.get('渠道对齐', '')).strip()
            
            fill_color = None
            if code_match == '否':
                fill_color = red_fill
            elif box_align == '是' and channel_align == '是':
                fill_color = green_fill
            elif box_align == '否' or channel_align == '否':
                fill_color = yellow_fill
            
            if idx in duplicate_indices:
                fill_color = orange_fill
            
            if fill_color:
                for col_idx in range(1, len(available_cols) + 1):
                    cell = worksheet.cell(row=row_num, column=col_idx).fill = fill_color
    
    print(f"已导出到 {filename}")


def export_backfill_to_original(original_file: str, compared_df: pd.DataFrame, output_filename: str):
    """
    导出回填结果 excel
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(original_file)
    ws = wb["包裹清单"]
    
    red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
    yellow_fill = PatternFill(start_color="EEFF00", end_color="EEFF00", fill_type="solid")
    green_fill = PatternFill(start_color="14E01E", end_color="14E01E", fill_type="solid")
    
    processed_count = 0
    filled_color_count = 0
    
    for idx, row in compared_df.iterrows():
        excel_row = row.get('_excel_row')
        col_scan = row.get('_excel_col_scan')
        col_damaged = row.get('_excel_col_damaged')
        
        if pd.isna(excel_row) or pd.isna(col_scan) or pd.isna(col_damaged):
            continue
        
        excel_row = int(excel_row)
        col_scan = int(col_scan)
        col_damaged = int(col_damaged)
        
        scan_box = row.get('扫描箱号', '')
        scan_channel = row.get('扫描渠道号', '')
        if pd.notna(scan_box) or pd.notna(scan_channel):
            box_str = str(scan_box) if pd.notna(scan_box) else ''
            channel_str = str(scan_channel) if pd.notna(scan_channel) else ''
            if box_str and channel_str:
                damaged_value = f"{box_str},{channel_str}"
            elif box_str:
                damaged_value = box_str
            elif channel_str:
                damaged_value = channel_str
            else:
                damaged_value = None
            
            if damaged_value:
                ws.cell(row=excel_row, column=col_damaged).value = damaged_value
        
        scan_ori = row.get('原始扫描序号', '')
        if pd.notna(scan_ori) and scan_ori:
            ws.cell(row=excel_row, column=col_scan).value = str(scan_ori)
        
        code_match = str(row.get('条码匹配', '')).strip()
        box_align = str(row.get('箱号对齐', '')).strip()
        channel_align = str(row.get('渠道对齐', '')).strip()
        
        has_scan_info = pd.notna(row.get('扫描箱号')) or pd.notna(row.get('扫描渠道号')) or pd.notna(row.get('原始扫描序号'))
        
        if not has_scan_info:
            fill_color = None
        elif code_match == '否':
            fill_color = red_fill
        elif box_align == '是' and channel_align == '是':
            fill_color = green_fill
        else:
            fill_color = yellow_fill
        
        if fill_color:
            ws.cell(row=excel_row, column=col_scan).fill = fill_color
            ws.cell(row=excel_row, column=col_damaged).fill = fill_color
            filled_color_count += 1
        
        processed_count += 1
    
    wb.save(output_filename)
    print(f"  回填统计: 处理 {processed_count} 行, 填充颜色 {filled_color_count} 行")
    print(f"已回填到 {output_filename}")


def compare_scan_to_pkg(df_scan: pd.DataFrame, df_pkg: pd.DataFrame) -> pd.DataFrame:
    """
    逆向比对
    """
    res = df_scan.copy()
    
    res["是否匹配"] = "否"
    res["预报单号"] = ""
    res["托盘序号"] = ""
    res["操作箱号"] = ""
    res["操作渠道号"] = ""
    
    pkg_lookup = []
    for idx, row in df_pkg.iterrows():
        tuo = row.get("托盘序号")
        pre = row.get("预报单号")
        key = f"{'' if pd.isna(tuo) else str(tuo).strip()}{'' if pd.isna(pre) else str(pre).strip()}"
        key = key.strip()
        if key:
            pkg_lookup.append({
                "key": key,
                "预报单号": pre,
                "托盘序号": tuo,
                "箱号": row.get("箱号"),
                "渠道号": row.get("渠道号")
            })

    for idx, row in res.iterrows():
        fba_code = row.get("fba条码")
        
        is_matched = False
        matched_pkg = None
        
        if pd.notna(fba_code) and str(fba_code).strip():
            fba_str = str(fba_code).strip()
            for item in pkg_lookup:
                if fba_str in item["key"]:
                    is_matched = True
                    matched_pkg = item
                    break
        
        if is_matched and matched_pkg:
            res.at[idx, "是否匹配"] = "是"
            res.at[idx, "预报单号"] = matched_pkg["预报单号"]
            res.at[idx, "托盘序号"] = matched_pkg["托盘序号"]
            res.at[idx, "操作箱号"] = matched_pkg["箱号"]
            res.at[idx, "操作渠道号"] = matched_pkg["渠道号"]
            
    return res


def export_unreport_with_colors(df: pd.DataFrame, filename: str):
    """
    导出未预报结果
    """
    scan_cols = [c for c in df.columns if c not in ["是否匹配", "预报单号", "托盘序号", "操作箱号", "操作渠道号", "_excel_row"]]
    if 'fba条码' in scan_cols:
        scan_cols.remove('fba条码')
        scan_cols.insert(scan_cols.index('条码') + 1 if '条码' in scan_cols else 0, 'fba条码')
        
    new_cols = ["是否匹配", "预报单号", "托盘序号", "操作箱号", "操作渠道号"]
    final_cols = scan_cols + new_cols
    final_cols = [c for c in final_cols if c in df.columns]
    
    export_df = df[final_cols].copy()
    export_df = export_df.reset_index(drop=True)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='Sheet1', index=False)
        worksheet = writer.sheets['Sheet1']
        
        red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
        yellow_fill = PatternFill(start_color="EEFF00", end_color="EEFF00", fill_type="solid")
        green_fill = PatternFill(start_color="14E01E", end_color="14E01E", fill_type="solid")
        
        for idx, row in export_df.iterrows():
            row_num = idx + 2
            is_matched = str(row.get("是否匹配", "")).strip()
            
            fill_color = None
            if is_matched == "否":
                fill_color = red_fill
            elif is_matched == "是":
                scan_box = str(row.get("箱号", "")).strip()
                op_box = str(row.get("操作箱号", "")).strip()
                scan_channel = str(row.get("渠道号", "")).strip()
                op_channel = str(row.get("操作渠道号", "")).strip()
                
                box_match = (scan_box == op_box) or (not scan_box and not op_box)
                channel_match = (scan_channel == op_channel) or (not scan_channel and not op_channel)
                
                if box_match and channel_match:
                    fill_color = green_fill
                else:
                    fill_color = yellow_fill
            
            if fill_color:
                for col_idx in range(1, len(final_cols) + 1):
                    cell = worksheet.cell(row=row_num, column=col_idx).fill = fill_color
                    
    print(f"已导出到 {filename}")


def filter_valid_boxes(df: pd.DataFrame) -> pd.DataFrame:
    """
    筛选逻辑：保留匹配的箱号
    """
    if df.empty or "箱号" not in df.columns or "是否匹配" not in df.columns:
        return df
        
    matched_rows = df[df["是否匹配"] == "是"]
    valid_boxes = set(matched_rows["箱号"].dropna().unique())
    filtered_df = df[df["箱号"].isin(valid_boxes)].copy()
    
    original_count = len(df)
    filtered_count = len(filtered_df)
    dropped_count = original_count - filtered_count
    if dropped_count > 0:
        print(f"筛选完成: 原数据 {original_count} 行 -> 筛选后 {filtered_count} 行 (过滤了 {dropped_count} 行未匹配箱号的数据)")
    else:
        print(f"筛选完成: 无数据被过滤 ({original_count} 行)")
        
    return filtered_df


def has_matches(df: pd.DataFrame) -> bool:
    if df.empty or "是否匹配" not in df.columns:
        return False
    return (df["是否匹配"] == "是").any()


def export_merged_with_colors(df_compare: pd.DataFrame, df_unreport: pd.DataFrame, filename: str):
    """
    导出合并结果：Sheet1=比较结果, Sheet2=未预报结果
    """
    # --- Sheet 1 ---
    required_cols_1 = ['预报单号', '托盘序号', '出库Ref', '破损/不可识别', '箱号', '渠道号', 
                     '条码匹配', '箱号对齐', '渠道对齐', '扫描箱号', '扫描渠道号', '扫描托盘号', '原始扫描序号']
    available_cols_1 = [col for col in required_cols_1 if col in df_compare.columns]
    export_df_1 = df_compare[available_cols_1].copy()
    export_df_1 = export_df_1.reset_index(drop=True)
    
    # --- Sheet 2 ---
    scan_cols = [c for c in df_unreport.columns if c not in ["是否匹配", "预报单号", "托盘序号", "操作箱号", "操作渠道号", "_excel_row"]]
    if 'fba条码' in scan_cols:
        scan_cols.remove('fba条码')
        scan_cols.insert(scan_cols.index('条码') + 1 if '条码' in scan_cols else 0, 'fba条码')
    new_cols = ["是否匹配", "预报单号", "托盘序号", "操作箱号", "操作渠道号"]
    final_cols_2 = scan_cols + new_cols
    final_cols_2 = [c for c in final_cols_2 if c in df_unreport.columns]
    
    export_df_2 = df_unreport[final_cols_2].copy()
    export_df_2 = export_df_2.reset_index(drop=True)
    
    red_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
    yellow_fill = PatternFill(start_color="EEFF00", end_color="EEFF00", fill_type="solid")
    green_fill = PatternFill(start_color="14E01E", end_color="14E01E", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFA500", fill_type="solid")

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1
        export_df_1.to_excel(writer, sheet_name='比较结果', index=False)
        ws1 = writer.sheets['比较结果']
        
        try:
            original_scan_col = export_df_1['原始扫描序号'].astype(str).apply(lambda x: x.strip())
            valid_mask = (original_scan_col != '') & (original_scan_col != 'nan') & (original_scan_col != 'None')
            valid_series = original_scan_col[valid_mask]
            duplicates = valid_series[valid_series.duplicated(keep=False)]
            duplicate_indices = set(duplicates.index)
        except Exception:
            duplicate_indices = set()

        for idx, row in export_df_1.iterrows():
            row_num = idx + 2
            code_match = str(row.get('条码匹配', '')).strip()
            box_align = str(row.get('箱号对齐', '')).strip()
            channel_align = str(row.get('渠道对齐', '')).strip()
            
            fill_color = None
            if code_match == '否':
                fill_color = red_fill
            elif box_align == '是' and channel_align == '是':
                fill_color = green_fill
            elif box_align == '否' or channel_align == '否':
                fill_color = yellow_fill
            if idx in duplicate_indices:
                fill_color = orange_fill
            
            if fill_color:
                for col_idx in range(1, len(available_cols_1) + 1):
                    ws1.cell(row=row_num, column=col_idx).fill = fill_color

        # Sheet 2
        export_df_2.to_excel(writer, sheet_name='未预报结果', index=False)
        ws2 = writer.sheets['未预报结果']
        
        for idx, row in export_df_2.iterrows():
            row_num = idx + 2
            is_matched = str(row.get("是否匹配", "")).strip()
            
            fill_color_2 = None
            if is_matched == "否":
                fill_color_2 = red_fill
            elif is_matched == "是":
                scan_box = str(row.get("箱号", "")).strip()
                op_box = str(row.get("操作箱号", "")).strip()
                scan_channel = str(row.get("渠道号", "")).strip()
                op_channel = str(row.get("操作渠道号", "")).strip()
                
                box_match = (scan_box == op_box) or (not scan_box and not op_box)
                channel_match = (scan_channel == op_channel) or (not scan_channel and not op_channel)
                
                if box_match and channel_match:
                    fill_color_2 = green_fill
                else:
                    fill_color_2 = yellow_fill
            
            if fill_color_2:
                for col_idx in range(1, len(final_cols_2) + 1):
                    ws2.cell(row=row_num, column=col_idx).fill = fill_color_2

    print(f"已导出合并报告到 {filename}")

# =================================================================================================
#  Main Logic
# =================================================================================================

def process_full_workflow(table_b_path: str, preprocessed_scan_df) -> bool:
    """
    执行完整流程：
    1. 预处理包裹清单 (Table B)
    2. 正向比对 (Pkg -> Scan)
    3. 逆向比对 (Scan -> Pkg) & 筛选
    4. 导出合并报告 (Sheet1=比较结果, Sheet2=未预报结果)
    5. 导出回填结果 (基于原始Excel格式回填)
    """
    table_b_path_obj = Path(table_b_path)
    table_b_name = table_b_path_obj.stem
    output_dir = Path("./compare_tables_test/output")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 定义输出文件名
    merged_report_file = output_dir / f"{table_b_name}_比较结果.xlsx"
    backfill_file = output_dir / f"回填结果_{table_b_name}.xlsx"
    
    # 强制覆盖模式 (可选: 检查并提示)
    if merged_report_file.exists() or backfill_file.exists():
        print(f"警告: 输出文件已存在，将被覆盖")
        
    # 1. 预处理包裹清单
    print(f"[{table_b_name}] 正在读取并预处理...")
    try:
        raw_pkg2 = preprocess_pkg_list(table_b_path)
        print(f"  预处理完成，共 {len(raw_pkg2)} 行数据")
    except Exception as e:
        print(f"  错误: 预处理失败 - {e}")
        return False
        
    # 2. 正向比对
    print(f"  正在执行正向比对 (比较结果)...")
    df_compare = compare_tables(preprocessed_scan_df, raw_pkg2)
    
    # 3. 逆向比对
    print(f"  正在执行逆向比对 (未预报结果)...")
    df_unreport = compare_scan_to_pkg(preprocessed_scan_df, raw_pkg2)
    
    # 筛选逆向结果
    df_unreport_filtered = filter_valid_boxes(df_unreport)
    print(f"  逆向结果筛选完毕: {len(df_unreport)} -> {len(df_unreport_filtered)} 行")

    # 4. 导出合并报告
    print(f"  正在导出合并报告: {merged_report_file.name}")
    export_merged_with_colors(df_compare, df_unreport_filtered, str(merged_report_file))
    
    # 5. 导出回填结果
    print(f"  正在导出回填结果: {backfill_file.name}")
    try:
        export_backfill_to_original(str(table_b_path_obj), df_compare, str(backfill_file))
    except Exception as e:
        print(f"  错误: 回填导出失败 - {e}")
    
    return True

def main():
    parser = argparse.ArgumentParser(description='主程序 v3: 生成合并比较报告 & 回填结果')
    parser.add_argument('table_a', nargs='?', 
                        default='./compare_tables_test/input_scan',
                        help='表A文件路径或文件夹(扫描数据)')
    parser.add_argument('table_b', nargs='?',
                        default='./compare_tables_test/input_pkg',
                        help='表B文件路径或文件夹(包裹清单)')
    
    args = parser.parse_args()
    
    # 1. 加载扫描数据
    table_a_path = Path(args.table_a)
    print(f"正在加载扫描数据: {table_a_path}")
    preprocessed_scan_df = load_scan_data(table_a_path)
    
    if preprocessed_scan_df is None:
        print("错误: 无法加载扫描数据，程序终止。")
        return

    # 2. 处理表B
    table_b_path = Path(args.table_b)
    
    if table_b_path.is_dir():
        print(f"\n检测到文件夹，开始批量处理: {table_b_path}")
        xlsx_files = sorted([f for f in table_b_path.glob("*.xlsx") if not f.name.startswith("~$")])
        if not xlsx_files:
            print(f"错误: 文件夹中没有找到 .xlsx 文件")
            return
        
        print(f"找到 {len(xlsx_files)} 个文件待处理...")
        processed_count = 0
        
        for idx, xlsx_file in enumerate(xlsx_files, 1):
            print(f"\n[{idx}/{len(xlsx_files)}] 处理文件: {xlsx_file.name}")
            print("-" * 60)
            result = process_full_workflow(str(xlsx_file), preprocessed_scan_df)
            if result: processed_count += 1
            
        print(f"\n批量处理完成！成功处理 {processed_count}/{len(xlsx_files)} 个文件。")
    else:
        if not table_b_path.exists():
            print(f"错误: 文件不存在 - {table_b_path}")
            return
        
        print(f"\n处理单文件: {table_b_path.name}")
        print("-" * 60)
        process_full_workflow(str(table_b_path), preprocessed_scan_df)
        print("\n处理完成。")

if __name__ == "__main__":
    main()
